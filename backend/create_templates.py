import os
import csv

def create_csv_template(dest_path):
    headers = [
        "nombre", "marca", "modelo", "categoria", "precio", "precio_anterior", 
        "tienda", "url_producto", "url_imagen", "aro", "talla", "material", 
        "transmision", "frenos", "suspension", "peso", "stock", "descripcion"
    ]
    
    samples = [
        [
            "Bicicleta Mountain Bike Marlin 5 Gen 3 Aro 29", "Trek", "Marlin 5 Gen 3", "bicicletas", 
            "449990", "549990", "Trek Chile", "https://www.trekbikes.com/cl/es_CL/marlin-5-gen-3/", 
            "https://images.unsplash.com/photo-1485965120184-e220f721d03e", "29", "M", "Aluminio", 
            "Shimano CUES 1x9", "Disco Hidráulico", "Delantera 100mm", "14.5", "1", 
            "Bicicleta de montaña de alto rendimiento ideal para senderos y uso urbano diario."
        ],
        [
            "Bicicleta Paseo Oxford Rally Aro 26", "Oxford", "Rally", "bicicletas", 
            "189990", "249990", "Oxford Store", "https://www.oxfordstore.cl/rally-aro-26.html", 
            "https://images.unsplash.com/photo-1485965120184-e220f721d03e", "26", "L", "Acero", 
            "Shimano Tourney 7v", "V-Brake", "Rígida", "15.2", "1", 
            "Bicicleta de paseo tradicional con tapabarros, parrilla y canasto frontal."
        ]
    ]
    
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    with open(dest_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(samples)
    print(f"[OK] Generated CSV template at {dest_path}")

def create_xlsx_template(dest_path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[WARN] openpyxl not installed. Attempting to install...")
        os.system("pip install openpyxl -q")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bicicletas Template"
    
    headers = [
        "nombre", "marca", "modelo", "categoria", "precio", "precio_anterior", 
        "tienda", "url_producto", "url_imagen", "aro", "talla", "material", 
        "transmision", "frenos", "suspension", "peso", "stock", "descripcion"
    ]
    
    samples = [
        [
            "Bicicleta Mountain Bike Marlin 5 Gen 3 Aro 29", "Trek", "Marlin 5 Gen 3", "bicicletas", 
            449990, 549990, "Trek Chile", "https://www.trekbikes.com/cl/es_CL/marlin-5-gen-3/", 
            "https://images.unsplash.com/photo-1485965120184-e220f721d03e", "29", "M", "Aluminio", 
            "Shimano CUES 1x9", "Disco Hidráulico", "Delantera 100mm", 14.5, 1, 
            "Bicicleta de montaña de alto rendimiento ideal para senderos y uso urbano diario."
        ],
        [
            "Bicicleta Paseo Oxford Rally Aro 26", "Oxford", "Rally", "bicicletas", 
            189990, 249990, "Oxford Store", "https://www.oxfordstore.cl/rally-aro-26.html", 
            "https://images.unsplash.com/photo-1485965120184-e220f721d03e", "26", "L", "Acero", 
            "Shimano Tourney 7v", "V-Brake", "Rígida", 15.2, 1, 
            "Bicicleta de paseo tradicional con tapabarros, parrilla y canasto frontal."
        ]
    ]
    
    # Write headers and format them
    font_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Dark blue
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font_bold
        cell.fill = fill_header
        cell.alignment = align_center
        
    # Write samples
    for row_idx, row_data in enumerate(samples, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
            
    # Auto-fit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    wb.save(dest_path)
    print(f"[OK] Generated XLSX template at {dest_path}")

if __name__ == "__main__":
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    imports_dir = os.path.join(project_root, "imports")
    
    create_csv_template(os.path.join(imports_dir, "bicicletas_template.csv"))
    create_xlsx_template(os.path.join(imports_dir, "bicicletas_template.xlsx"))
